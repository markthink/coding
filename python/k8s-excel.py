# -*- coding: utf-8 -*-

import os,subprocess,codecs;
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import  Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

# 定义翻译函数
class KubernetesDoc:

    # 首次全量翻译/增量翻译更新一下路径地址
    # baseurl = "/Users/dxwsker/works/k8smeetup/kubernetes.github.io/docs"
    baseurl = "/Users/dxwsker/works/k8smeetup/k8s-update/docs"

    # 获取所有需要翻译的 MD 文件列表
    def mdfiles(self,path):
        current_files = os.listdir(path)
        all_files = []
        for file_name in current_files:
            full_file_name = os.path.join(path, file_name)
            if (os.path.splitext(full_file_name)[1] == '.md' and os.path.dirname(full_file_name).find("node_modules") == -1):
                # print(os.path.dirname(full_file_name))
                all_files.append(full_file_name)

            if (os.path.splitext(full_file_name)[1] == '.html' and os.path.dirname(full_file_name).find("node_modules") == -1):
                all_files.append(full_file_name)

            if os.path.isdir(full_file_name):
                next_level_files = self.mdfiles(full_file_name)
                all_files.extend(next_level_files)

        return all_files

if __name__ == '__main__':

    k8s = KubernetesDoc()

    MDfiles = k8s.mdfiles(k8s.baseurl)

    wb = Workbook()
    ws = wb.active

    # 设置表头：
    header = ['Files','Should translate?','Already exist','Translator','Our translation', 'Chinese Version']
    initset=1
    for c in range(0, len(header)):
        ws.cell(column=initset+c, row=1, value = header[c])

    # 设置表头样式
    YELLOWFill = PatternFill(start_color='FFF68F', end_color='FFB90F', fill_type='solid')
    for row in ws.iter_rows(min_row=1, max_col=6, max_row=1):
        for cell in row:
            cell.font = Font(size=15)
            cell.fill = YELLOWFill

    ws.column_dimensions[get_column_letter(1)].width = 100
    ws.column_dimensions[get_column_letter(2)].width = 25
    ws.column_dimensions[get_column_letter(3)].width = 25
    ws.column_dimensions[get_column_letter(4)].width = 30
    ws.column_dimensions[get_column_letter(5)].width = 50
    ws.column_dimensions[get_column_letter(6)].width = 50


    # 填充内容 - 定义表头偏移
    offset = 2
    for i in range(0, len(MDfiles)):
        # ws.cell(column=1, row=offset + i, value=MDfiles[i].split('kubernetes.github.io')[1])
        ws.cell(column=1, row=offset + i, value=MDfiles[i].split('k8s-update')[1])
        ws.cell(column=2, row=offset + i, value=1)
        ws.cell(column=3, row=offset + i, value=0)
    wb.save("k8s-2007-08-15.xlsx")

